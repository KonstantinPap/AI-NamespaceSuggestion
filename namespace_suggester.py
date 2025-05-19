import os
import re
import csv
from typing import Dict, List, Tuple, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm

from langchain_community.chat_models import AzureChatOpenAI
from langchain.schema import SystemMessage, HumanMessage

HC_ROOT = "C:/Repos/DevOps/HC-Work/Product_MED/Product_MED_AL/app/"
MTC_ROOT = "C:/Repos/DevOps/MTC-Work/Product_MED_Tech365/Product_MED_Tech/app/"
ANALYZE_ROOTS = [HC_ROOT, MTC_ROOT]

SEARCH_ROOTS = [
    HC_ROOT,
    MTC_ROOT,
    "C:/Repos/Github/StefanMaron/MSDyn365BC.Code.History/BaseApp/Source/Base Application/",
    "C:/Repos/DevOps/HC-Work/Product_KBA/Product_KBA_BC_AL/app/",
]
CSV_OUTPUT = "namespace_suggestions.csv"
# Azure OpenAI Konfiguration
OPENAI_API_KEY = os.environ.get("AZURE_OPENAI_KEY") or os.environ.get("OPENAI_API_KEY")
OPENAI_API_BASE = os.environ.get("AZURE_OPENAI_ENDPOINT")
OPENAI_API_VERSION = os.environ.get("AZURE_OPENAI_API_VERSION", "2024-12-01-preview")
OPENAI_DEPLOYMENT = os.environ.get("AZURE_OPENAI_DEPLOYMENT", "gpt-4o-mini")


OBJECT_PATTERN = re.compile(r'^(table|page|codeunit|report|xmlport|query|enum|interface|controladdin|pageextension|tableextension|enumextension|profile|dotnet|entitlement|permissionset|permissionsetextension|reportextension|enumvalue|entitlementset|entitlementsetextension)\s+(\d+)?\s*"?([\w\d_]+)"?', re.IGNORECASE)
NAMESPACE_PATTERN = re.compile(r'(?:Namespace\s*=\s*"([\w\d_.]+)"|namespace\s+([\w\d_.]+)\s*;)', re.IGNORECASE)
REF_PATTERN = re.compile(
    r'(?:'
        r'(Database|Table|Page|Codeunit|Report|XmlPort|Query|Enum)\s*::\s*"?([\w\d_]+)"?'
        r'|'
        r':\s*(Record|Page|Codeunit|Report|XmlPort|Query|Enum)\s+("?[\w\d_]+"?)'
    r')',
    re.IGNORECASE
)

# Namespace-Beschreibungen wie in namespace_review.py
allowed_namespaces_with_desc = [
        ("Microsoft", "Microsoft Standardfunktionalität und Basiskomponenten"),
        ("Microsoft.API", "Microsoft API-spezifische Komponenten und Schnittstellen"),
        ("Microsoft.API.Upgrade", "Upgrade-bezogene APIs und Migrationshilfen"),
        ("Microsoft.API.Webhooks", "Webhooks-Integration und Ereignisbenachrichtigungen"),
        ("Microsoft.AccountantPortal", "Funktionen für das Accountant Portal"),
        ("Microsoft.Assembly.Comment", "Kommentare und Anmerkungen im Bereich Montage"),
        ("Microsoft.Assembly.Costing", "Kalkulation und Kostenrechnung für Montage"),
        ("Microsoft.Assembly.Document", "Dokumentenmanagement im Montagebereich"),
        ("Microsoft.Assembly.History", "Historie und Protokollierung von Montageprozessen"),
        ("Microsoft.Assembly.Posting", "Buchungen und Verbuchungen im Montagebereich"),
        ("Microsoft.Assembly.Reports", "Berichte und Auswertungen zur Montage"),
        ("Microsoft.Assembly.Setup", "Einrichtung und Konfiguration der Montage"),
        ("Microsoft.Bank.BankAccount", "Bankkontenverwaltung"),
        ("Microsoft.Bank.Check", "Scheckverwaltung und -verarbeitung"),
        ("Microsoft.Bank.Deposit", "Einzahlungen und Bankeinlagen"),
        ("Microsoft.Bank.DirectDebit", "Lastschriftverfahren und SEPA-Lastschriften"),
        ("Microsoft.Bank.Ledger", "Bankbuchhaltung und Konten"),
        ("Microsoft.Bank.Payment", "Zahlungsabwicklung und Zahlungsverkehr"),
        ("Microsoft.Bank.PositivePay", "Positive Pay-Funktionen für Banken"),
        ("Microsoft.Bank.Reconciliation", "Bankabstimmung und Kontenabgleich"),
        ("Microsoft.Bank.Reports", "Bankbezogene Berichte und Auswertungen"),
        ("Microsoft.Bank.Setup", "Einrichtung und Konfiguration von Bankfunktionen"),
        ("Microsoft.Bank.Statement", "Bankauszüge und Kontoauszugsverarbeitung"),
        ("Microsoft.Booking", "Buchungsfunktionen und Reservierungen"),
        ("Microsoft.CRM.Analysis", "CRM-Analysen und Auswertungen"),
        ("Microsoft.CRM.BusinessRelation", "Geschäftsbeziehungen im CRM"),
        ("Microsoft.CRM.Campaign", "Kampagnenmanagement im CRM"),
        ("Microsoft.CRM.Comment", "Kommentare und Notizen im CRM"),
        ("Microsoft.CRM.Contact", "Kontaktverwaltung im CRM"),
        ("Microsoft.CRM.Duplicates", "Duplikaterkennung und -management im CRM"),
        ("Microsoft.CRM.Interaction", "Interaktionen und Aktivitäten im CRM"),
        ("Microsoft.CRM.Opportunity", "Vertriebschancen und Opportunities im CRM"),
        ("Microsoft.CRM.Outlook", "Outlook-Integration für CRM"),
        ("Microsoft.CRM.Profiling", "Profiling und Segmentierung im CRM"),
        ("Microsoft.CRM.Reports", "CRM-Berichte und Auswertungen"),
        ("Microsoft.CRM.RoleCenters", "Rollencenter für CRM-Anwender"),
        ("Microsoft.CRM.Segment", "Segmentierung und Zielgruppen im CRM"),
        ("Microsoft.CRM.Setup", "Einrichtung und Konfiguration des CRM"),
        ("Microsoft.CRM.Task", "Aufgabenmanagement im CRM"),
        ("Microsoft.CRM.Team", "Teamverwaltung im CRM"),
        ("Microsoft.CashFlow.Account", "Liquiditätskonten für Cashflow-Planung"),
        ("Microsoft.CashFlow.Comment", "Kommentare im Bereich Cashflow"),
        ("Microsoft.CashFlow.Forecast", "Cashflow-Prognosen und -Planung"),
        ("Microsoft.CashFlow.Reports", "Berichte zur Liquiditätsplanung"),
        ("Microsoft.CashFlow.Setup", "Einrichtung der Cashflow-Funktionen"),
        ("Microsoft.CashFlow.Worksheet", "Arbeitsblätter für Cashflow-Analysen"),
        ("Microsoft.CostAccounting.Account", "Kostenrechnungskonten"),
        ("Microsoft.CostAccounting.Allocation", "Kostenverteilung und Umlagen"),
        ("Microsoft.CostAccounting.Budget", "Kostenrechnungsbudgets"),
        ("Microsoft.CostAccounting.Journal", "Kostenrechnungsjournale"),
        ("Microsoft.CostAccounting.Ledger", "Kostenrechnungshauptbuch"),
        ("Microsoft.CostAccounting.Posting", "Buchungen in der Kostenrechnung"),
        ("Microsoft.CostAccounting.Reports", "Berichte zur Kostenrechnung"),
        ("Microsoft.CostAccounting.Setup", "Einrichtung der Kostenrechnung"),
        ("Microsoft.EServices.EDocument", "Elektronische Dokumente und eServices"),
        ("Microsoft.Finance.AllocationAccount", "Verteilungskonten im Finanzwesen"),
        ("Microsoft.Finance.AllocationAccount.Purchase", "Verteilungskonten für Einkauf"),
        ("Microsoft.Finance.AllocationAccount.Sales", "Verteilungskonten für Verkauf"),
        ("Microsoft.Finance.Analysis", "Finanzanalysen und Auswertungen"),
        ("Microsoft.Finance.AuditFileExport", "Export von Audit-Dateien"),
        ("Microsoft.Finance.Consolidation", "Konsolidierung im Finanzbereich"),
        ("Microsoft.Finance.Currency", "Währungsmanagement"),
        ("Microsoft.Finance.Deferral", "Abgrenzungen und Rechnungsabgrenzungsposten"),
        ("Microsoft.Finance.Dimension", "Dimensionen im Finanzwesen"),
        ("Microsoft.Finance.Dimension.Correction", "Korrekturen von Dimensionen"),
        ("Microsoft.Finance.FinancialReports", "Finanzberichte und Auswertungen"),
        ("Microsoft.Finance.GeneralLedger.Account", "Sachkonten im Hauptbuch"),
        ("Microsoft.Finance.GeneralLedger.Budget", "Budgets im Hauptbuch"),
        ("Microsoft.Finance.GeneralLedger.Journal", "Journale im Hauptbuch"),
        ("Microsoft.Finance.GeneralLedger.Ledger", "Hauptbuchfunktionen"),
        ("Microsoft.Finance.GeneralLedger.Posting", "Buchungen im Hauptbuch"),
        ("Microsoft.Finance.GeneralLedger.Preview", "Vorschau von Hauptbuchbuchungen"),
        ("Microsoft.Finance.GeneralLedger.Reports", "Berichte zum Hauptbuch"),
        ("Microsoft.Finance.GeneralLedger.Reversal", "Stornierungen im Hauptbuch"),
        ("Microsoft.Finance.GeneralLedger.Setup", "Einrichtung des Hauptbuchs"),
        ("Microsoft.Finance.Payroll", "Lohn- und Gehaltsabrechnung"),
        ("Microsoft.Finance.ReceivablesPayables", "Debitoren- und Kreditorenbuchhaltung"),
        ("Microsoft.Finance.RoleCenters", "Rollencenter für Finanzanwender"),
        ("Microsoft.Finance.SalesTax", "Umsatzsteuerverwaltung"),
        ("Microsoft.Finance.VAT", "Mehrwertsteuerverwaltung"),
        ("Microsoft.Finance.VAT.Calculation", "Berechnung der Mehrwertsteuer"),
        ("Microsoft.Finance.VAT.Clause", "Mehrwertsteuerklauseln"),
        ("Microsoft.Finance.VAT.Ledger", "Mehrwertsteuerhauptbuch"),
        ("Microsoft.Finance.VAT.RateChange", "Mehrwertsteuersatzänderungen"),
        ("Microsoft.Finance.VAT.Registration", "Mehrwertsteuerregistrierung"),
        ("Microsoft.Finance.VAT.Reporting", "Mehrwertsteuerberichte"),
        ("Microsoft.Finance.VAT.Setup", "Einrichtung der Mehrwertsteuer"),
        ("Microsoft.FixedAssets.Depreciation", "Abschreibungen auf Anlagegüter"),
        ("Microsoft.FixedAssets.FixedAsset", "Verwaltung von Anlagegütern"),
        ("Microsoft.FixedAssets.Insurance", "Versicherung von Anlagegütern"),
        ("Microsoft.FixedAssets.Journal", "Anlagenjournale"),
        ("Microsoft.FixedAssets.Ledger", "Anlagenhauptbuch"),
        ("Microsoft.FixedAssets.Maintenance", "Wartung von Anlagegütern"),
        ("Microsoft.FixedAssets.Posting", "Buchungen im Anlagenbereich"),
        ("Microsoft.FixedAssets.Reports", "Berichte zu Anlagegütern"),
        ("Microsoft.FixedAssets.Setup", "Einrichtung der Anlagenbuchhaltung"),
        ("Microsoft.Foundation.Address", "Adressverwaltung"),
        ("Microsoft.Foundation.Attachment", "Anhänge und Dokumentenmanagement"),
        ("Microsoft.Foundation.AuditCodes", "Audit-Codes und Prüfungsfunktionen"),
        ("Microsoft.Foundation.BatchProcessing", "Stapelverarbeitung und Hintergrundprozesse"),
        ("Microsoft.Foundation.Calendar", "Kalenderfunktionen"),
        ("Microsoft.Foundation.Comment", "Kommentare und Notizen"),
        ("Microsoft.Foundation.Company", "Unternehmensverwaltung"),
        ("Microsoft.Foundation.Enums", "Aufzählungstypen und Enums"),
        ("Microsoft.Foundation.ExtendedText", "Erweiterte Textfunktionen"),
        ("Microsoft.Foundation.Navigate", "Navigationsfunktionen"),
        ("Microsoft.Foundation.NoSeries", "Nummernserienverwaltung"),
        ("Microsoft.Foundation.PaymentTerms", "Zahlungsbedingungen"),
        ("Microsoft.Foundation.Period", "Periodenverwaltung"),
        ("Microsoft.Foundation.Reporting", "Berichtswesen und Reporting"),
        ("Microsoft.Foundation.Shipping", "Versand und Logistik"),
        ("Microsoft.Foundation.Task", "Aufgabenverwaltung"),
        ("Microsoft.Foundation.UOM", "Mengeneinheitenverwaltung"),
        ("Microsoft.HumanResources.Absence", "Abwesenheitsverwaltung"),
        ("Microsoft.HumanResources.Analysis", "Analysen im Personalbereich"),
        ("Microsoft.HumanResources.Comment", "Kommentare im Personalbereich"),
        ("Microsoft.HumanResources.Employee", "Mitarbeiterverwaltung"),
        ("Microsoft.HumanResources.Payables", "Verbindlichkeiten im Personalbereich"),
        ("Microsoft.HumanResources.Reports", "Berichte im Personalbereich"),
        ("Microsoft.HumanResources.RoleCenters", "Rollencenter für Personalwesen"),
        ("Microsoft.HumanResources.Setup", "Einrichtung des Personalbereichs"),
        ("Microsoft.Integration.D365Sales", "Integration mit Dynamics 365 Sales"),
        ("Microsoft.Integration.Dataverse", "Integration mit Dataverse"),
        ("Microsoft.Integration.Entity", "Integration von Entitäten"),
        ("Microsoft.Integration.FieldService", "Integration mit Field Service"),
        ("Microsoft.Integration.Graph", "Microsoft Graph-Integration"),
        ("Microsoft.Integration.PowerBI", "Power BI-Integration"),
        ("Microsoft.Integration.SyncEngine", "Synchronisationsengine"),
        ("Microsoft.Intercompany", "Intercompany-Funktionen"),
        ("Microsoft.Intercompany.BankAccount", "Intercompany-Bankkonten"),
        ("Microsoft.Intercompany.Comment", "Kommentare im Intercompany-Bereich"),
        ("Microsoft.Intercompany.DataExchange", "Datenaustausch zwischen Unternehmen"),
        ("Microsoft.Intercompany.Dimension", "Dimensionen im Intercompany-Bereich"),
        ("Microsoft.Intercompany.GLAccount", "Sachkonten im Intercompany-Bereich"),
        ("Microsoft.Intercompany.Inbox", "Eingangskorb für Intercompany"),
        ("Microsoft.Intercompany.Journal", "Intercompany-Journale"),
        ("Microsoft.Intercompany.Outbox", "Ausgangskorb für Intercompany"),
        ("Microsoft.Intercompany.Partner", "Intercompany-Partner"),
        ("Microsoft.Intercompany.Reports", "Berichte im Intercompany-Bereich"),
        ("Microsoft.Intercompany.Setup", "Einrichtung des Intercompany-Bereichs"),
        ("Microsoft.Inventory", "Bestandsverwaltung"),
        ("Microsoft.Inventory.Analysis", "Bestandsanalysen"),
        ("Microsoft.Inventory.Availability", "Bestandsverfügbarkeit"),
        ("Microsoft.Inventory.BOM", "Stücklistenverwaltung"),
        ("Microsoft.Inventory.BOM.Tree", "Stücklistenstruktur"),
        ("Microsoft.Inventory.Comment", "Kommentare im Bestandsbereich"),
        ("Microsoft.Inventory.Costing", "Kostenrechnung im Bestandsbereich"),
        ("Microsoft.Inventory.Costing.ActionMessage", "Handlungsempfehlungen zur Kostenrechnung"),
        ("Microsoft.Inventory.Counting", "Inventurzählung"),
        ("Microsoft.Inventory.Counting.Comment", "Kommentare zur Inventurzählung"),
        ("Microsoft.Inventory.Counting.Document", "Dokumente zur Inventurzählung"),
        ("Microsoft.Inventory.Counting.History", "Historie der Inventurzählung"),
        ("Microsoft.Inventory.Counting.Journal", "Inventurzählungsjournale"),
        ("Microsoft.Inventory.Counting.Recording", "Erfassung der Inventurzählung"),
        ("Microsoft.Inventory.Counting.Reports", "Berichte zur Inventurzählung"),
        ("Microsoft.Inventory.Counting.Tracking", "Nachverfolgung der Inventurzählung"),
        ("Microsoft.Inventory.Document", "Bestandsdokumente"),
        ("Microsoft.Inventory.History", "Bestandshistorie"),
        ("Microsoft.Inventory.Intrastat", "Intrastat-Meldungen"),
        ("Microsoft.Inventory.Item", "Artikelverwaltung"),
        ("Microsoft.Inventory.Item.Attribute", "Artikelattribute"),
        ("Microsoft.Inventory.Item.Catalog", "Artikelkatalog"),
        ("Microsoft.Inventory.Item.Picture", "Artikelbilder"),
        ("Microsoft.Inventory.Item.Substitution", "Artikelersatz"),
        ("Microsoft.Inventory.Journal", "Bestandsjournale"),
        ("Microsoft.Inventory.Ledger", "Bestandshauptbuch"),
        ("Microsoft.Inventory.Location", "Lagerorte"),
        ("Microsoft.Inventory.MarketingText", "Marketingtexte für Artikel"),
        ("Microsoft.Inventory.Planning", "Bestandsplanung"),
        ("Microsoft.Inventory.Posting", "Buchungen im Bestandsbereich"),
        ("Microsoft.Inventory.Reconciliation", "Bestandsabstimmung"),
        ("Microsoft.Inventory.Reports", "Berichte zur Bestandsverwaltung"),
        ("Microsoft.Inventory.Requisition", "Bestellanforderungen"),
        ("Microsoft.Inventory.RoleCenters", "Rollencenter für Bestandsverwaltung"),
        ("Microsoft.Inventory.Setup", "Einrichtung der Bestandsverwaltung"),
        ("Microsoft.Inventory.StandardCost", "Standardkosten im Bestand"),
        ("Microsoft.Inventory.Tracking", "Bestandsnachverfolgung"),
        ("Microsoft.Inventory.Transfer", "Bestandsübertragungen"),
        ("Microsoft.Iventory.Item", "Artikelverwaltung (Tippfehler: Inventory)"),
        ("Microsoft.Manufacturing.Capacity", "Kapazitätsplanung in der Fertigung"),
        ("Microsoft.Manufacturing.Comment", "Kommentare im Fertigungsbereich"),
        ("Microsoft.Manufacturing.Document", "Fertigungsdokumente"),
        ("Microsoft.Manufacturing.Family", "Fertigungsfamilien"),
        ("Microsoft.Manufacturing.Forecast", "Fertigungsprognosen"),
        ("Microsoft.Manufacturing.Integration", "Integrationen im Fertigungsbereich"),
        ("Microsoft.Manufacturing.Journal", "Fertigungsjournale"),
        ("Microsoft.Manufacturing.MachineCenter", "Maschinenzentren in der Fertigung"),
        ("Microsoft.Manufacturing.Planning", "Fertigungsplanung"),
        ("Microsoft.Manufacturing.ProductionBOM", "Produktionsstücklisten"),
        ("Microsoft.Manufacturing.Reports", "Berichte zur Fertigung"),
        ("Microsoft.Manufacturing.RoleCenters", "Rollencenter für Fertigung"),
        ("Microsoft.Manufacturing.Routing", "Arbeitspläne in der Fertigung"),
        ("Microsoft.Manufacturing.Setup", "Einrichtung der Fertigung"),
        ("Microsoft.Manufacturing.StandardCost", "Standardkosten in der Fertigung"),
        ("Microsoft.Manufacturing.WorkCenter", "Arbeitszentren in der Fertigung"),
        ("Microsoft.Pricing.Asset", "Preisfindung für Anlagen"),
        ("Microsoft.Pricing.Calculation", "Preisberechnung"),
        ("Microsoft.Pricing.PriceList", "Preislistenverwaltung"),
        ("Microsoft.Pricing.Reports", "Berichte zur Preisfindung"),
        ("Microsoft.Pricing.Source", "Preisquellen"),
        ("Microsoft.Pricing.Worksheet", "Arbeitsblätter zur Preisfindung"),
        ("Microsoft.Projects.Project.Analysis", "Projektanalysen"),
        ("Microsoft.Projects.Project.Archive", "Projektarchivierung"),
        ("Microsoft.Projects.Project.Job", "Projektaufträge"),
        ("Microsoft.Projects.Project.Journal", "Projektjournale"),
        ("Microsoft.Projects.Project.Ledger", "Projekthauptbuch"),
        ("Microsoft.Projects.Project.Planning", "Projektplanung"),
        ("Microsoft.Projects.Project.Posting", "Projektbuchungen"),
        ("Microsoft.Projects.Project.Pricing", "Projektpreisfindung"),
        ("Microsoft.Projects.Project.Reports", "Projektberichte"),
        ("Microsoft.Projects.Project.Setup", "Einrichtung des Projektbereichs"),
        ("Microsoft.Projects.Project.WIP", "Work in Progress im Projektbereich"),
        ("Microsoft.Projects.Resources.Analysis", "Analyse von Projektressourcen"),
        ("Microsoft.Projects.Resources.Journal", "Journale für Projektressourcen"),
        ("Microsoft.Projects.Resources.Ledger", "Hauptbuch für Projektressourcen"),
        ("Microsoft.Projects.Resources.Pricing", "Preisfindung für Projektressourcen"),
        ("Microsoft.Projects.Resources.Reports", "Berichte zu Projektressourcen"),
        ("Microsoft.Projects.Resources.Resource", "Projektressourcenverwaltung"),
        ("Microsoft.Projects.Resources.Setup", "Einrichtung der Projektressourcen"),
        ("Microsoft.Projects.RoleCenters", "Rollencenter für Projekte"),
        ("Microsoft.Projects.TimeSheet", "Projekt-Zeiterfassung"),
        ("Microsoft.Purchases.Analysis", "Einkaufsanalysen"),
        ("Microsoft.Purchases.Archive", "Archivierung von Einkaufsbelegen"),
        ("Microsoft.Purchases.Comment", "Kommentare im Einkaufsbereich"),
        ("Microsoft.Purchases.Document", "Einkaufsdokumente"),
        ("Microsoft.Purchases.History", "Einkaufshistorie"),
        ("Microsoft.Purchases.Payables", "Verbindlichkeiten im Einkauf"),
        ("Microsoft.Purchases.Posting", "Buchungen im Einkauf"),
        ("Microsoft.Purchases.Pricing", "Preisfindung im Einkauf"),
        ("Microsoft.Purchases.Remittance", "Zahlungsavis im Einkauf"),
        ("Microsoft.Purchases.Reports", "Berichte zum Einkauf"),
        ("Microsoft.Purchases.RoleCenters", "Rollencenter für Einkauf"),
        ("Microsoft.Purchases.Setup", "Einrichtung des Einkaufsbereichs"),
        ("Microsoft.Purchases.Vendor", "Lieferantenverwaltung"),
        ("Microsoft.RoleCenters", "Allgemeine Rollencenter"),
        ("Microsoft.Sales.Analysis", "Verkaufsanalysen"),
        ("Microsoft.Sales.Archive", "Archivierung von Verkaufsbelegen"),
        ("Microsoft.Sales.Comment", "Kommentare im Verkaufsbereich"),
        ("Microsoft.Sales.Customer", "Kundenverwaltung"),
        ("Microsoft.Sales.Document", "Verkaufsdokumente"),
        ("Microsoft.Sales.FinanceCharge", "Finanzierungsgebühren im Verkauf"),
        ("Microsoft.Sales.History", "Verkaufshistorie"),
        ("Microsoft.Sales.Peppol", "PEPPOL-Integration im Verkauf"),
        ("Microsoft.Sales.Posting", "Buchungen im Verkauf"),
        ("Microsoft.Sales.Pricing", "Preisfindung im Verkauf"),
        ("Microsoft.Sales.Receivables", "Forderungen aus Lieferungen und Leistungen"),
        ("Microsoft.Sales.Reminder", "Zahlungserinnerungen im Verkauf"),
        ("Microsoft.Sales.Reports", "Berichte zum Verkauf"),
        ("Microsoft.Sales.RoleCenters", "Rollencenter für Verkauf"),
        ("Microsoft.Sales.Setup", "Einrichtung des Verkaufsbereichs"),
        ("Microsoft.Service.Analysis", "Serviceanalysen"),
        ("Microsoft.Service.Archive", "Archivierung von Servicebelegen"),
        ("Microsoft.Service.BaseApp", "Service-Basisfunktionen"),
        ("Microsoft.Service.CashFlow", "Servicebezogene Cashflow-Funktionen"),
        ("Microsoft.Service.Comment", "Kommentare im Servicebereich"),
        ("Microsoft.Service.Contract", "Serviceverträge"),
        ("Microsoft.Service.Customer", "Servicekundenverwaltung"),
        ("Microsoft.Service.Document", "Servicedokumente"),
        ("Microsoft.Service.Email", "E-Mail-Kommunikation im Service"),
        ("Microsoft.Service.History", "Servicehistorie"),
        ("Microsoft.Service.Item", "Serviceartikelverwaltung"),
        ("Microsoft.Service.Ledger", "Servicehauptbuch"),
        ("Microsoft.Service.Loaner", "Leihstellungen im Service"),
        ("Microsoft.Service.Maintenance", "Wartung im Servicebereich"),
        ("Microsoft.Service.Posting", "Buchungen im Servicebereich"),
        ("Microsoft.Service.Pricing", "Preisfindung im Service"),
        ("Microsoft.Service.Reports", "Berichte zum Service"),
        ("Microsoft.Service.Resources", "Servicebezogene Ressourcen"),
        ("Microsoft.Service.RoleCenters", "Rollencenter für Service"),
        ("Microsoft.Service.Setup", "Einrichtung des Servicebereichs"),
        ("Microsoft.Shared.Report", "Geteilte Berichte und Reportings"),
        ("Microsoft.System.Threading", "Nebenläufigkeit und Threading"),
        ("Microsoft.Upgrade", "Upgrade- und Migrationsfunktionen"),
        ("Microsoft.Utilities", "Hilfsfunktionen und Utilities"),
        ("Microsoft.Warehouse.ADCS", "Automatisierte Datenerfassung im Lager"),
        ("Microsoft.Warehouse.Activity", "Lageraktivitäten"),
        ("Microsoft.Warehouse.Activity.History", "Historie der Lageraktivitäten"),
        ("Microsoft.Warehouse.Availability", "Lagerverfügbarkeit"),
        ("Microsoft.Warehouse.Comment", "Kommentare im Lagerbereich"),
        ("Microsoft.Warehouse.CrossDock", "Cross-Docking im Lager"),
        ("Microsoft.Warehouse.Document", "Lagerdokumente"),
        ("Microsoft.Warehouse.History", "Lagerhistorie"),
        ("Microsoft.Warehouse.InternalDocument", "Interne Lagerdokumente"),
        ("Microsoft.Warehouse.InventoryDocument", "Bestandsdokumente im Lager"),
        ("Microsoft.Warehouse.Journal", "Lagerjournale"),
        ("Microsoft.Warehouse.Ledger", "Lagerhauptbuch"),
        ("Microsoft.Warehouse.Posting", "Buchungen im Lagerbereich"),
        ("Microsoft.Warehouse.Reports", "Berichte zum Lager"),
        ("Microsoft.Warehouse.Request", "Lageranforderungen"),
        ("Microsoft.Warehouse.RoleCenters", "Rollencenter für Lager"),
        ("Microsoft.Warehouse.Setup", "Einrichtung des Lagerbereichs"),
        ("Microsoft.Warehouse.Structure", "Lagerstruktur"),
        ("Microsoft.Warehouse.Tracking", "Lagerverfolgung"),
        ("Microsoft.Warehouse.Worksheet", "Lagerarbeitsblätter"),
        ("Microsoft.costaccounting.Reports", "Berichte zur Kostenrechnung (Legacy)"),
        ("Microsoft.eServices.OnlineMap", "Online-Kartendienste"),
        ("System.AI", "Künstliche Intelligenz und Machine Learning"),
        ("System.Apps", "Systemanwendungen und App-Management"),
        ("System.Automation", "Automatisierungsfunktionen"),
        ("System.Azure.Identity", "Azure-Identitätsdienste"),
        ("System.DataAdministration", "Datenadministration und -management"),
        ("System.DateTime", "Datum- und Zeitfunktionen"),
        ("System.Device", "Geräteverwaltung"),
        ("System.Diagnostics", "Diagnose- und Überwachungsfunktionen"),
        ("System.EMail", "E-Mail-Kommunikation"),
        ("System.Email", "E-Mail-Kommunikation"),
        ("System.Environment", "Systemumgebung und Konfiguration"),
        ("System.Environment.Configuration", "Konfiguration der Systemumgebung"),
        ("System.Feedback", "Feedback- und Rückmeldungsfunktionen"),
        ("System.Globalization", "Globalisierung und Lokalisierung"),
        ("System.IO", "Datei- und Datenzugriff"),
        ("System.Integration", "Systemintegration"),
        ("System.Integration.PowerBI", "Power BI-Integration auf Systemebene"),
        ("System.Media", "Medienverwaltung"),
        ("System.Privacy", "Datenschutz und Privatsphäre"),
        ("System.Reflection", "Reflexion und Metadaten"),
        ("System.Security.AccessControl", "Zugriffssteuerung und Sicherheit"),
        ("System.Security.Authentication", "Authentifizierungsfunktionen"),
        ("System.Security.Encryption", "Verschlüsselung und Sicherheit"),
        ("System.Security.User", "Benutzerverwaltung und -sicherheit"),
        ("System.Telemetry", "Telemetrie und Überwachung"),
        ("System.TestTools", "Testwerkzeuge und Testautomatisierung"),
        ("System.TestTools.CodeCoverage", "Testabdeckung und Coverage"),
        ("System.TestTools.TestRunner", "Testausführung"),
        ("System.Text", "Textverarbeitung"),
        ("System.Threading", "Nebenläufigkeit und Threading"),
        ("System.Tooling", "Entwicklungswerkzeuge"),
        ("System.Utilities", "Systemnahe Hilfsfunktionen"),
        ("System.Visualization", "Visualisierung und Darstellung"),
        ("System.Xml", "XML-Verarbeitung"),
        # KUMAVISION Base (KBA)
        ("UDI", "Unique Device Identification (KUMAVISION base/KBA)"),
        ("Call", "Servicetickets und Anrufe (KUMAVISION base/KBA)"),
        ("LIF", "Etikettenmanagement/handling (KUMAVISION base/KBA)"),
        ("OrderQuote", "Angebots- und Auftragsverwaltung (KUMAVISION base/KBA)"),
        ("InventorySummary", "Bestandsübersicht (KUMAVISION base/KBA)"),
        ("Common", "KUMAVISION Basiskomponenten. Root Namespace"),
        # HC/MTC-spezifisch       
        ("ECE", "Elektronischer Datenaustausch (HC/MTC)"),
        ("MDR", "Medical Device Regulation (HC/MTC)"),
        ("Common", "Gemeinsame Komponenten für MTC/HC"),
]

allowed_namespaces = [ns for ns, desc in allowed_namespaces_with_desc]

def find_al_files(root: str) -> List[str]:
    al_files = []
    for dirpath, _, filenames in os.walk(root):
        for f in filenames:
            if f.lower().endswith(".al"):
                al_files.append(os.path.join(dirpath, f))
    return al_files

def extract_object_info(filepath: str):
    with open(filepath, encoding="utf-8") as f:
        lines = f.readlines()
    obj_type, obj_name, namespace = None, None, None
    for line in lines:
        if not obj_type:
            m = OBJECT_PATTERN.match(line.strip())
            if m:
                obj_type = m.group(1)
                obj_name = m.group(3)
        if not namespace:
            n = NAMESPACE_PATTERN.search(line)
            if n:
                namespace = n.group(1) or n.group(2)
        if obj_type and obj_name and namespace is not None:
            break
    al_code = "".join(lines)
    return obj_type, obj_name, namespace, al_code

def extract_references(al_code: str) -> List[str]:
    return list(set(m[1] for m in REF_PATTERN.findall(al_code)))

def index_al_objects(roots: List[str]) -> Dict[str, Dict]:
    """Indexiere alle AL-Objekte nach Name (case-insensitive)."""
    index = {}
    for root in roots:
        for filepath in find_al_files(root):
            try:
                obj_type, obj_name, namespace, al_code = extract_object_info(filepath)
                if obj_type and obj_name:
                    key = obj_name.lower()
                    index[key] = {
                        "object_type": obj_type,
                        "object_name": obj_name,
                        "namespace": namespace,
                        "filepath": filepath,
                        "al_code": al_code
                    }
            except Exception:
                continue
    return index

def index_al_objects_with_type_and_name(roots: List[str]) -> Dict[Tuple[str, str], Dict]:
    """
    Indexiere alle AL-Objekte nach (object_type.lower(), object_name.lower()).
    Liefert Dict mit Typ und Name als Schlüssel, damit Referenzen besser aufgelöst werden können.
    Fortschrittsanzeige mit tqdm.
    """
    index = {}
    all_files = []
    for root in roots:
        all_files.extend(find_al_files(root))
    for filepath in tqdm(all_files, desc="Indexiere AL-Objekte", unit="Datei"):
        try:
            obj_type, obj_name, namespace, al_code = extract_object_info(filepath)
            if obj_type and obj_name:
                key = (obj_type.lower(), obj_name.lower())
                index[key] = {
                    "object_type": obj_type,
                    "object_name": obj_name,
                    "namespace": namespace,
                    "filepath": filepath,
                    "al_code": al_code
                }
        except Exception:
            continue
    return index

def extract_reference_tuples(al_code: str) -> List[Tuple[str, str]]:
    """
    Extrahiere Referenzen als (object_type, object_name)-Tupel aus dem AL-Code.
    """
    refs = []
    for m in REF_PATTERN.findall(al_code):
        # m[0] oder m[2] ist der Typ, m[1] oder m[3] ist der Name
        if m[1]:
            refs.append((m[0].lower(), m[1].lower()))
        elif m[3]:
            refs.append((m[2].lower(), m[3].replace('"', '').lower()))
    return list(set(refs))

# "Besonders wichtig ist, wenn Objekte auf Einrichtuungen (Setup-Tabellen) verweisen"
        # TODO Verweise auf MED Einrichtungen
        # TODO Berückischtige schon bisher gemachte Entscheidungen
        # TODO MEDTEC genause gewichten wie HC? Falls abweichender Namespace, dann in die Vorschlag, kommasepariert aufnehmen
        # TODO Cll Objekt Call gilt nur für HC, für die MTC ist das service
        # TODO BinCode MTC
        # TODO HC und MTC gesondert behandeln

def suggest_namespace_llm(obj_info, ref_infos):
    prompt = (
        "Du bist ein Experte für Microsoft Dynamics 365 Business Central AL-Entwicklung und die Vergabe von Namespaces.\n"
        "Analysiere das folgende AL-Objekt und schlage einen passenden Namespace vor. "
        "Beziehe dich dabei auf die Namenskonventionen der Microsoft Base Application. " 
        "Wenn im Standard (Base Application) für ein Objekt oder dessen Funktionalität bereits ein Namespace wie z.B. 'System', 'Sales', etc. verwendet wird, "
        "sollen die HC- und MTC-Objekte möglichst denselben Namespace verwenden. "
        "Die Entscheidung für den Namespace soll sich vorrangig an den Objekten der Base Application orientieren, aber auch die Namespace ECE und MDR sollten mit einkalkuliert werden.\n"        
        "Falls ein anderer Namespace sinnvoller ist, begründe dies nachvollziehbar.\n"
        "Die Begründung (\"reason\") und alle Alternativen im JSON-Output müssen ausschließlich auf DEUTSCH formuliert sein.\n"
        "Du darfst ausschließlich einen Namespace aus folgender Liste verwenden (keinen anderen):\n"
        "WICHTIG: Gib im Feld \"namespace\" ausschließlich den Kurznamen (z.B. 'EServices', 'EDocuments', 'Finance', 'Inventory', etc.) aus der Liste an – NICHT den vollständigen Namespace-Pfad wie 'Microsoft.EServices.EDocument' oder 'System.Environment.Configuration'!\n"
        "Wenn ein Namespace-Pfad wie 'System.Environment.Configuration' in der Liste steht, darfst du nur einen einzelnen Teil daraus wählen, z.B. entweder 'Environment' oder 'Configuration', aber niemals den gesamten Pfad oder mehrere Teile kombiniert.\n"
        "Beachte außerdem: Der Suffix 'SI' im Objektnamen steht für 'Single Instance' und der Suffix 'Sub' steht im Normalfall für eine Subscriber-Codeunit. "
        "Berücksichtige diese Bedeutungen bei deiner Analyse und Empfehlung.\n"
        "Beispiel für den Output:\n"
        '{"namespace": "Environment", "reason": "...", "alternatives": [{"namespace": "Configuration", "reason": "..."}]}'
    )
    prompt += "\n".join([f"- {ns}: {desc}" for ns, desc in allowed_namespaces_with_desc])
    prompt += (
        "\nFalls keiner dieser Namespaces fachlich passt, wähle 'Custom' und begründe dies ausführlich.\n"
        f"\nObjekttyp: {obj_info['object_type']}\n"
        f"Objektname: {obj_info['object_name']}\n"
        f"AL-Code:\n{obj_info['al_code']}\n"
    )
    if ref_infos:
        prompt += "\nKontext zu referenzierten Objekten:\n"
        for ref in ref_infos:
            prompt += f"- Name: {ref.get('object_name','')}, Namespace: {ref.get('namespace','')}\n"
    prompt += (
        "\nDeine Aufgabe:\n"
        "- Analysiere, ob und wie das Objekt oder ähnliche Objekte in der Base Application einem bestimmten Namespace zugeordnet sind.\n"
        "- Schlage einen passenden Namespace aus der Liste der erlaubten Namespaces vor (bevorzugt den Namespace der Base Application, falls vorhanden).\n"
        "- Begründe deine Entscheidung ausführlich in deutscher Sprache.\n"
        "- Schlage, falls sinnvoll, alternative Namespaces vor und begründe diese Alternativen in deutscher Sprache.\n"
        "Gib das Ergebnis als JSON im folgenden Format zurück (ALLE Begründungen auf DEUTSCH!):\n"
        '{"namespace": "...", "reason": "...", "alternatives": [{"namespace": "...", "reason": "..."}]}'
        "Deine Empfehlung:"
    )
    llm = AzureChatOpenAI(
        openai_api_key=OPENAI_API_KEY,
        azure_endpoint=OPENAI_API_BASE,
        openai_api_version=OPENAI_API_VERSION,
        deployment_name=OPENAI_DEPLOYMENT,
        temperature=0.5,
        max_tokens=800,
    )
    messages = [
        SystemMessage(content="Du bist ein erfahrener AL-Entwickler und Namespace-Experte."),
        HumanMessage(content=prompt)
    ]
    try:
        response = llm.invoke(messages)
        import json as pyjson
        text = response.content
        match = re.search(r'\{.*\}', text, re.DOTALL)
        if match:
            data = pyjson.loads(match.group(0))
            ns = data.get("namespace", "")
            reason = data.get("reason", "")
            alternatives = []
            for alt in data.get("alternatives", []):
                alternatives.append((alt.get("namespace", ""), alt.get("reason", "")))
            return ns, reason, alternatives, text
        else:
            return "", "Konnte kein JSON aus Azure OpenAI-Antwort extrahieren.", [], text
    except Exception as e:
        return "", f"Azure OpenAI-Fehler: {e}", [], ""

HC_PREFIX = "KVSMED"
MTC_PREFIX = "KVSMTC"

def remove_prefix(name: str, prefix: str) -> str:
    if name.upper().startswith(prefix.upper()):
        return name[len(prefix):]
    return name

def build_hc_mtc_object_map(analyze_obj_index):
    """
    Gruppiert Objekte aus HC und MTC nach Typ und Namen ohne Prefix.
    Gibt Dict mit key = (object_type, name_ohne_prefix) und value = {"hc": {...}, "mtc": {...}}
    """
    grouped = {}
    for (otype, oname), obj in analyze_obj_index.items():
        if oname.upper().startswith(HC_PREFIX):
            name_noprefix = remove_prefix(oname, HC_PREFIX)
            key = (otype, name_noprefix.lower())
            grouped.setdefault(key, {})["hc"] = obj
        elif oname.upper().startswith(MTC_PREFIX):
            name_noprefix = remove_prefix(oname, MTC_PREFIX)
            key = (otype, name_noprefix.lower())
            grouped.setdefault(key, {})["mtc"] = obj
        else:
            # Falls kein Prefix, trotzdem aufnehmen (z.B. für Basiskomponenten)
            key = (otype, oname.lower())
            grouped.setdefault(key, {})["hc"] = obj  # Default zu HC
    return grouped

def read_existing_csv(csv_path: str) -> set:
    """
    Lese bereits analysierte Objekte aus der CSV (ObjectType, HC ObjectName, MTC ObjectName als Schlüssel).
    """
    done = set()
    if not os.path.exists(csv_path):
        return done
    with open(csv_path, newline='', encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            key = (
                row.get("ObjectType", "").strip().lower(),
                row.get("HC ObjectName", "").strip().lower(),
                row.get("MTC ObjectName", "").strip().lower()
            )
            done.add(key)
    return done

import time
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def write_results_to_excel(rows, fieldnames, excel_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Namespace Vorschläge"
    ws.append(fieldnames)
    for row in rows:
        ws.append([row.get(col, "") for col in fieldnames])
    # Formatierung: Zeilenumbruch für lange Textspalten
    wrap_cols = ["Namespace Begründung", "Alternative Namespace Begründung", "Analyse"]
    for idx, col in enumerate(fieldnames, 1):
        if col in wrap_cols:
            for cell in ws[get_column_letter(idx)]:
                cell.alignment = Alignment(wrap_text=True)
    # Optional: Spaltenbreite anpassen
    for idx, col in enumerate(fieldnames, 1):
        ws.column_dimensions[get_column_letter(idx)].width = max(20, len(col) + 2)
    wb.save(excel_path)

def is_obsolete_object(al_code: str) -> bool:
    """
    Prüft, ob das gesamte Objekt als obsolete markiert ist.
    Es werden nur Objekte als obsolete erkannt, wenn das Obsolete-Attribut direkt im Objektkopf steht,
    nicht wenn nur Felder, Prozeduren oder andere Member obsolete sind.
    """
    # Suche nach [Obsolete(...)] direkt in den ersten Zeilen (Objektkopf)
    # Ignoriere Zeilen mit Feld-/Methodendefinitionen
    header_lines = "\n".join(al_code.splitlines()[:10])
    return (
        re.search(r'^\s*\[.*Obsolete.*\]', header_lines, re.IGNORECASE | re.MULTILINE)
        or re.search(r'^\s*Obsolete\s*=\s*.*;', header_lines, re.IGNORECASE | re.MULTILINE)
    ) is not None

def get_context_key(obj_info):
    """
    Liefert einen Schlüssel für den Kontext, z.B. für Setup-Tabellen oder Objekte mit gleichem Namensstamm.
    """
    # Beispiel: Gruppiere nach object_type + Namensstamm (ohne Suffix wie SI/Sub)
    name = obj_info["object_name"]
    # Entferne Suffixe wie SI, Sub
    name_base = re.sub(r'(SI|Sub)$', '', name, flags=re.IGNORECASE)
    return (obj_info["object_type"].lower(), name_base.lower())

def get_forced_namespace(obj_info, solution_type):
    """
    Gibt einen Namespace zurück, falls durch Namensbestandteile oder Obsolete-Status erzwungen.
    """
    name = obj_info["object_name"]
    al_code = obj_info["al_code"]
    if is_obsolete_object(al_code):
        return "Obsolete"
    if "MDR" in name:
        return "MDR"
    if "Cll" in name:
        return "Call" if solution_type == "HC" else "Service"
    return None

def main():
    # Index für Referenz-Kontext (alle Roots)
    ref_obj_index = index_al_objects_with_type_and_name(SEARCH_ROOTS)
    # Index für zu analysierende Objekte (nur HC/MTC)
    analyze_obj_index = index_al_objects_with_type_and_name(ANALYZE_ROOTS)

    # Alle Objekte einzeln betrachten, HC/MTC explizit unterscheiden
    all_objs = []
    for (otype, oname), obj in analyze_obj_index.items():
        solution_type = None
        if oname.upper().startswith(HC_PREFIX):
            solution_type = "HC"
        elif oname.upper().startswith(MTC_PREFIX):
            solution_type = "MTC"
        else:
            # Default zu HC, falls nicht eindeutig
            solution_type = "HC"
        all_objs.append((solution_type, obj))

    fieldnames = [
        "ObjectType",
        "SolutionType",
        "ObjectName",
        "Namespace Vorschlag",
        "Namespace Begründung",
        "Alternative Namespace Vorschlag",
        "Alternative Namespace Begründung",
        "Dateipfad",
        "Analyse"
    ]

    already_done = read_existing_csv(CSV_OUTPUT)
    results = []
    context_map = {}  # context_key -> list of obj_idx in results

    # 1. Alle Objekte analysieren und Namespace-Vorschlag im Speicher sammeln
    for idx, (solution_type, obj_info) in enumerate(tqdm(all_objs, desc="Namespace-Vorschläge", unit="Objekt"), 1):
        otype = obj_info["object_type"]
        obj_name = obj_info["object_name"]
        key = (otype, solution_type, obj_name.lower(),)
        if key in already_done:
            continue

        # Forced Namespace durch Regeln
        forced_ns = get_forced_namespace(obj_info, solution_type)
        ns, reason, alternatives, analyse = None, "", [], ""
        if forced_ns:
            ns = forced_ns
            reason = f"Namespace durch Regel erzwungen: {forced_ns}"
            analyse = reason
        else:
            # Referenzen analysieren (mit Typ und Name, Kontext aus ref_obj_index)
            ref_infos = []
            if obj_info["al_code"]:
                for ref_type, ref_name in extract_reference_tuples(obj_info["al_code"]):
                    ref_obj = ref_obj_index.get((ref_type, ref_name))
                    if ref_obj:
                        ref_infos.append({"object_type": ref_obj["object_type"], "object_name": ref_obj["object_name"], "namespace": ref_obj["namespace"]})
            ns, reason, alternatives, analyse = suggest_namespace_llm(obj_info, ref_infos)

        alt_ns = "; ".join([a[0] for a in alternatives])
        alt_reason = "; ".join([a[1] for a in alternatives])
        row = {
            "ObjectType": otype,
            "SolutionType": solution_type,
            "ObjectName": obj_name,
            "Namespace Vorschlag": ns,
            "Namespace Begründung": reason,
            "Alternative Namespace Vorschlag": alt_ns,
            "Alternative Namespace Begründung": alt_reason,
            "Dateipfad": obj_info["filepath"],
            "Analyse": analyse.strip().replace(chr(10), ' ')
        }
        results.append(row)
        # Kontext-Mapping für spätere Konsistenzprüfung
        ctx_key = get_context_key(obj_info)
        context_map.setdefault(ctx_key, []).append(len(results) - 1)

    # 2. Nachbearbeitung: Konsistenz im Kontext sicherstellen
    for ctx_key, idx_list in context_map.items():
        # Prüfe, ob alle Namespace-Vorschläge im Kontext gleich sind
        ns_counter = {}
        for idx in idx_list:
            ns = results[idx]["Namespace Vorschlag"]
            if ns:
                ns_counter[ns] = ns_counter.get(ns, 0) + 1
        if ns_counter:
            # Wähle den häufigsten Namespace im Kontext
            best_ns = max(ns_counter.items(), key=lambda x: x[1])[0]
            for idx in idx_list:
                # Nur überschreiben, wenn nicht forced (z.B. Obsolete, MDR, Cll)
                if not get_forced_namespace({"object_name": results[idx]["ObjectName"], "al_code": ""}, results[idx]["SolutionType"]):
                    results[idx]["Namespace Vorschlag"] = best_ns
                    results[idx]["Namespace Begründung"] += f" (Konsistenz im Kontext: {best_ns})"

    # 3. Schreibe Ergebnisse in CSV/Excel
    write_header = not os.path.exists(CSV_OUTPUT) or os.stat(CSV_OUTPUT).st_size == 0
    with open(CSV_OUTPUT, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        for row in results:
            writer.writerow(row)
    excel_path = CSV_OUTPUT.replace(".csv", ".xlsx")
    write_results_to_excel(results, fieldnames, excel_path)

if __name__ == "__main__":
    main()
